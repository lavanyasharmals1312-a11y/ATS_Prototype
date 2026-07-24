import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from uuid import UUID
from typing import Optional

from app.dependencies import get_current_user, get_candidate_service
from app.services.candidate_service import CandidateService
from app.schemas.candidate import (
    CandidateFilters, CandidateUpdate,
    CandidateResponse, CandidateListItem
)
from app.models.user import User
from app.services.file_service import build_file_response
from app.config import settings

router = APIRouter(prefix="/api/v1", tags=["candidates"])


@router.get("/candidates")
async def list_candidates(
    search: Optional[str] = Query(None),
    skills: Optional[str] = Query(None, description="Comma-separated: Python,React"),
    location: Optional[str] = Query(None),
    experience_min: Optional[float] = Query(None, ge=0),
    experience_max: Optional[float] = Query(None, le=50),
    notice_period: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    service: CandidateService = Depends(get_candidate_service),
    current_user: User = Depends(get_current_user),
):
    skills_list = [s.strip() for s in skills.split(",")] if skills else None
    filters = CandidateFilters(
        search=search, skills=skills_list, location=location,
        experience_min=experience_min, experience_max=experience_max,
        notice_period=notice_period, source=source,
    )
    result = await service.list_candidates(filters, page, page_size)
    result["items"] = [CandidateListItem.model_validate(c) for c in result["items"]]
    return {"success": True, "data": result}


@router.get("/candidates/export/excel")
async def export_candidates_excel(
    search: Optional[str] = Query(None),
    skills: Optional[str] = Query(None, description="Comma-separated: Python,React"),
    location: Optional[str] = Query(None),
    experience_min: Optional[float] = Query(None, ge=0),
    experience_max: Optional[float] = Query(None, le=50),
    notice_period: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    service: CandidateService = Depends(get_candidate_service),
    current_user: User = Depends(get_current_user),
):
    """Export all matching candidates as a styled Excel file."""
    skills_list = [s.strip() for s in skills.split(",")] if skills else None
    filters = CandidateFilters(
        search=search, skills=skills_list, location=location,
        experience_min=experience_min, experience_max=experience_max,
        notice_period=notice_period, source=source,
    )
    # Fetch all (no pagination) — up to 5000 rows
    result = await service.list_candidates(filters, page=1, page_size=5000)
    candidates = result["items"]

    excel_bytes = _build_excel(candidates)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"candidates_{timestamp}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_excel(candidates) -> bytes:
    """Build a styled xlsx workbook from a list of candidate ORM objects."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    # ── Header style ────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")  # dark navy
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill("solid", fgColor="EEF2F7")  # light blue-grey for alt rows

    # ── Columns ─────────────────────────────────────────────────────────
    columns = [
        ("Name",                 22),
        ("Email",                28),
        ("Phone",                15),
        ("Designation",          22),
        ("Company",              22),
        ("Experience (Yrs)",     16),
        ("Current Location",     18),
        ("Preferred Location",   18),
        ("Notice Period",        14),
        ("Current CTC",          14),
        ("Expected CTC",         14),
        ("Highest Qualification",18),
        ("University",           24),
        ("Skills",               40),
        ("Source",               12),
        ("Duplicate Status",     16),
        ("Added On",             16),
    ]

    ws.row_dimensions[1].height = 32
    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ────────────────────────────────────────────────────────
    for row_idx, c in enumerate(candidates, 2):
        is_alt = (row_idx % 2 == 0)
        row_fill = alt_fill if is_alt else PatternFill()

        skills_str = ", ".join(c.skills) if c.skills else ""
        added_on = (
            c.created_at.strftime("%Y-%m-%d") if c.created_at else ""
        )
        exp = float(c.experience_years) if c.experience_years is not None else ""

        row_values = [
            c.candidate_name or "",
            c.candidate_email or "",
            c.candidate_phone or "",
            c.current_designation or "",
            c.current_company or "",
            exp,
            c.current_location or "",
            c.preferred_location or "",
            c.notice_period or "",
            c.current_ctc or "",
            c.expected_ctc or "",
            c.highest_qualification or "",
            c.university or "",
            skills_str,
            c.source or "",
            c.duplicate_status or "",
            added_on,
        ]

        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 14))
            if is_alt:
                cell.fill = row_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/candidates/{candidate_id}")
async def get_candidate(
    candidate_id: UUID,
    service: CandidateService = Depends(get_candidate_service),
    current_user: User = Depends(get_current_user),
):
    candidate = await service.get_candidate(candidate_id)
    return {"success": True, "data": CandidateResponse.model_validate(candidate)}


@router.patch("/candidates/{candidate_id}")
async def update_candidate(
    candidate_id: UUID,
    data: CandidateUpdate,
    service: CandidateService = Depends(get_candidate_service),
    current_user: User = Depends(get_current_user),
):
    candidate = await service.update_candidate(candidate_id, data)
    return {"success": True, "data": CandidateResponse.model_validate(candidate)}


@router.delete("/candidates/{candidate_id}")
async def deactivate_candidate(
    candidate_id: UUID,
    service: CandidateService = Depends(get_candidate_service),
    current_user: User = Depends(get_current_user),
):
    await service.deactivate_candidate(candidate_id)
    return {"success": True, "message": "Candidate deactivated"}


@router.get("/candidates/{candidate_id}/resume")
async def download_resume(
    candidate_id: UUID,
    service: CandidateService = Depends(get_candidate_service),
    current_user: User = Depends(get_current_user),
):
    candidate = await service.get_candidate(candidate_id)
    if not candidate.resumes:
        from fastapi import HTTPException
        raise HTTPException(404, "No resume found for this candidate")
    latest = sorted(candidate.resumes, key=lambda r: r.uploaded_at, reverse=True)[0]
    return build_file_response(latest.file_path, latest.original_filename, settings.upload_dir)