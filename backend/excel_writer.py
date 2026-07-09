import os
from openpyxl import Workbook, load_workbook

# Save Excel inside the backend folder
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "Master_Tracker_Gemini.xlsx")

HEADERS = [
    "S.No",
    "Record ID",
    "Duplicate Status",
    "Alt ID",
    "Name",
    "Contact No",
    "Email",
    "Level",
    "Current Organization",
    "Total Experience",
    "Current CTC",
    "Expected CTC",
    "Notice Period",
    "Current Location",
    "Preferred Location",
    "Offer in Hand",
    "Reason for Job Change",
    "Highest Qualification",
    "University",
    "Communication Rating",
    "Comments",
    "Shared On",
    "Source Name",
    "Role",
    "Skills",
    "HR Name"
]


def append_candidate(candidate):

    # Create Excel if it doesn't exist
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        ws.append(HEADERS)

    # Serial Number
    sno = ws.max_row

    # Convert skills list to comma-separated string
    skills = candidate.get("skills", [])

    if isinstance(skills, list):
        skills = ", ".join(skills)

    ws.append([
        sno,                                       # S.No
        "",                                        # Record ID
        "",                                        # Duplicate Status
        "",                                        # Alt ID
        candidate.get("candidate_name", ""),       # Name
        candidate.get("candidate_phone", ""),      # Contact No
        candidate.get("candidate_email", ""),      # Email
        "",                                        # Level
        candidate.get("current_company", ""),      # Current Organization
        candidate.get("experience_years", ""),     # Total Experience
        candidate.get("current_ctc", ""),          # Current CTC
        candidate.get("expected_ctc", ""),         # Expected CTC
        candidate.get("notice_period", ""),        # Notice Period
        candidate.get("current_location", ""),     # Current Location
        candidate.get("preferred_location", ""),   # Preferred Location
        "",                                        # Offer in Hand
        "",                                        # Reason for Job Change
        candidate.get("highest_qualification", ""),# Highest Qualification
        candidate.get("university", ""),           # University
        "",                                        # Communication Rating
        "",                                        # Comments
        "",                                        # Shared On
        "",                                        # Source Name
        candidate.get("current_designation", ""),  # Role
        skills,                                    # Skills
        ""                                         # HR Name
    ])

    wb.save(EXCEL_FILE)

    print(f"Excel saved at: {EXCEL_FILE}")

    return EXCEL_FILE